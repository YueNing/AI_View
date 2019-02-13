import os,sys
import openpyxl
sys.path.append('../mk')
sys.path.append('./mk')
os.environ.setdefault("DJANGO_SETTINGS_MODULE", 'mk.settings')
import django
django.setup()
from backend.models import Movies, Movies_Shot, k_Shot

word2vector = {'drama':
                    {
                        'loss':['death','get up', 'discussion'], 
                        'kidnapping':['car ride', 'phone call', 'dramatic reunion'], 
                        'identity crisis':['despair','resolution','knowledge moment']
                    },
                'action':
                    {
                        'war':['attack', 'climax', 'make a plan'],
                        'space':['no return', 'aliens attack', 'missing your family'],
                        'competition':['win', 'lose', 'car chase']
                    },
                'comedy':
                    {
                        'confusion':['a object', 'a person'],
                        'surrealism':['shapes', 'colors'],
                        'relationship':['get to know', 'forced situation']
                    }
                }

def search_genre(plot):
    """
    >>> search_genre(['get up', 'discussion'])
    ['drama', 'drama']
    """
    genre = ['drama', 'drama']
    return genre

def search_theme(plot):
    """
    >>> search_theme(['get up', 'discussion'])
    ['loss', 'loss']
    """
    plot = ['loss', 'loss']
    return plot

def get_plot():
    plot = []
    return plot

def get_id():
    id = 1
    return id

def split_to_list(s):
    return s.split(',')

def main(filexl):
    wb = openpyxl.load_workbook(filexl)
    ws = wb.get_sheet_by_name(wb.sheetnames[0])
    first_row = True
    for row in ws.iter_rows():
        if first_row:
            first_row = False
            continue
        shot_id = row[0].value
        genres = split_to_list(row[1].value)
        themes = split_to_list(row[2].value)
        plots = split_to_list(row[3].value)
        duration = row[4].value
        if not k_Shot.objects.filter(shot_id=shot_id):
            k_shot = k_Shot(shot_id=shot_id, url='http://127.0.0.1/my_video_scenes_tmp/videos/{}'.format(shot_id+'.mp4'))
            k_shot.setgenres(genres)
            k_shot.setthemes(themes)
            k_shot.setplots(plots)
            k_shot.setduration(duration)
            k_shot.save()
            print("save {}".format(shot_id))
        else:
            k_shot = k_Shot(id=k_Shot.objects.filter(shot_id=shot_id)[0].id, shot_id=shot_id, url='http://127.0.0.1/my_video_scenes_tmp/videos/{}'.format(shot_id+'.mp4'))
            k_shot.setgenres(genres)
            k_shot.setthemes(themes)
            k_shot.setplots(plots)
            k_shot.setduration(duration)
            k_shot.save()
            print('update {}'.format(shot_id))

if __name__=='__main__':
    main('data_ai_view.xlsx')
    # import doctest
    # doctest.testmod(verbose=True)